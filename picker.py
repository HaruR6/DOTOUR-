from __future__ import annotations

import argparse
import random
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


DEFAULT_EXCEL = "ドトールメニュー商品価格一覧_2026-07-30(1).xlsx"
SHEET_NAME = "商品一覧"


@dataclass(frozen=True)
class Product:
    name: str
    genre: str
    size: str
    price: int

    @property
    def label(self) -> str:
        size = "" if self.size in ("", "—", "-") else f"（{self.size}）"
        return f"{self.name}{size}"


DRINK_GENRES = {"ホットドリンク", "アイスドリンク"}
FOOD50_GENRES = {"ミラノサンド", "モーニング・セット"}
CAKE_GENRES = {"デザート"}


def calculate_discount(products: Iterable[Product]) -> tuple[int, list[str]]:
    """
    セット割引額と内訳を返す。

    ルール:
    - ドリンク1点につき、割引は最大1回だけ。
    - ドリンク＋ミラノサンドは1組につき50円引き。
    - まだ割引に使われていないドリンク＋ケーキは1組につき30円引き。
    - 割引額が大きい50円対象フードとの組合せを優先する。
    """
    products_list = list(products)

    drink_count = sum(product.genre in DRINK_GENRES for product in products_list)
    food50_count = sum(product.genre in FOOD50_GENRES for product in products_list)
    cake_count = sum(product.genre in CAKE_GENRES for product in products_list)

    food50_pairs = min(drink_count, food50_count)
    remaining_drinks = drink_count - food50_pairs
    cake_pairs = min(remaining_drinks, cake_count)

    food50_discount = food50_pairs * 50
    cake_discount = cake_pairs * 30
    discount = food50_discount + cake_discount

    details: list[str] = []
    if food50_pairs:
        details.append(
            f"ドリンク＋50円対象フード割引: -{food50_discount:,}円"
            f"（{food50_pairs}組）"
        )
    if cake_pairs:
        details.append(
            f"ドリンク＋ケーキ割引: -{cake_discount:,}円"
            f"（{cake_pairs}組）"
        )

    return discount, details


def payable_total(products: Iterable[Product]) -> int:
    products_list = list(products)
    subtotal = sum(product.price for product in products_list)
    discount, _ = calculate_discount(products_list)
    return subtotal - discount


def load_products(excel_path: Path) -> list[Product]:
    """Excelの商品一覧シートを読み込む。"""
    if not excel_path.exists():
        raise FileNotFoundError(f"Excelファイルが見つかりません: {excel_path}")

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"シート『{SHEET_NAME}』が見つかりません。")

    sheet = workbook[SHEET_NAME]
    headers = {
        str(cell.value).strip(): index
        for index, cell in enumerate(next(sheet.iter_rows()), start=0)
        if cell.value is not None
    }

    required = ["商品名", "ジャンル", "サイズ", "価格（税込・円）"]
    missing = [name for name in required if name not in headers]
    if missing:
        raise ValueError(f"必要な列がありません: {', '.join(missing)}")

    products: list[Product] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = row[headers["商品名"]]
        genre = row[headers["ジャンル"]]
        size = row[headers["サイズ"]]
        price = row[headers["価格（税込・円）"]]

        if name is None or genre is None or price is None:
            continue

        try:
            numeric_price = int(price)
        except (TypeError, ValueError):
            continue

        products.append(
            Product(
                name=str(name).strip(),
                genre=str(genre).strip(),
                size="—" if size is None else str(size).strip(),
                price=numeric_price,
            )
        )

    return products


def available_genres(products: Iterable[Product]) -> list[str]:
    return sorted({product.genre for product in products})


def pick_products(
    products: list[Product],
    genre_counts: dict[str, int],
    budget: int,
    rng: random.Random,
) -> list[Product]:
    """ジャンルごとの個数と割引後予算を満たす組合せを返す。"""
    if budget < 0:
        raise ValueError("上限金額は0円以上にしてください。")
    if not genre_counts or sum(genre_counts.values()) <= 0:
        raise ValueError("少なくとも1個以上の商品を指定してください。")

    grouped: dict[str, dict[str, list[Product]]] = {}
    for genre, count in genre_counts.items():
        if count < 0:
            raise ValueError("個数は0以上にしてください。")
        if count == 0:
            continue
        genre_products = [p for p in products if p.genre == genre]
        if not genre_products:
            raise ValueError(f"ジャンル『{genre}』の商品がありません。")
        by_name: dict[str, list[Product]] = {}
        for product in genre_products:
            by_name.setdefault(product.name, []).append(product)
        if len(by_name) < count:
            raise ValueError(
                f"ジャンル『{genre}』には、重複なしで選べる商品が"
                f"{len(by_name)}種類しかありません。"
            )
        for variants in by_name.values():
            rng.shuffle(variants)
        grouped[genre] = by_name

    genre_order = [g for g, c in genre_counts.items() if c > 0]
    rng.shuffle(genre_order)
    total_requested = sum(genre_counts.values())

    def search_genre(genre_index: int, selected: list[Product], used_names: set[str]) -> list[Product] | None:
        if genre_index == len(genre_order):
            return selected.copy() if payable_total(selected) <= budget else None
        genre = genre_order[genre_index]
        required = genre_counts[genre]
        names = list(grouped[genre].keys())
        rng.shuffle(names)

        def choose(start_index: int, remaining: int, local_selected: list[Product], local_names: set[str]) -> list[Product] | None:
            if remaining == 0:
                return search_genre(genre_index + 1, selected + local_selected, used_names | local_names)
            if len(names) - start_index < remaining:
                return None
            for i in range(start_index, len(names)):
                name = names[i]
                if name in used_names or name in local_names:
                    continue
                for product in grouped[genre][name]:
                    trial = selected + local_selected + [product]
                    if sum(p.price for p in trial) - 50 * total_requested > budget:
                        continue
                    result = choose(i + 1, remaining - 1, local_selected + [product], local_names | {name})
                    if result is not None:
                        return result
            return None
        return choose(0, required, [], set())

    result = search_genre(0, [], set())
    if result is None:
        raise ValueError("指定したジャンル別個数と予算を同時に満たす組合せが見つかりませんでした。")
    rng.shuffle(result)
    return result


def ask_int(message: str, minimum: int = 1) -> int:
    while True:
        text = input(message).strip().replace(",", "")
        try:
            number = int(text)
        except ValueError:
            print("整数で入力してください。")
            continue
        if number < minimum:
            print(f"{minimum}以上の整数を入力してください。")
            continue
        return number


def interactive(products: list[Product], rng: random.Random) -> None:
    genres = available_genres(products)
    print("\n利用できるジャンル")
    for index, genre in enumerate(genres, start=1):
        print(f"  {index}. {genre}")

    while True:
        selected = input(
            "\nジャンル名または番号をカンマ区切りで入力してください\n"
            "例: 2,8 または アイスドリンク,デザート\n> "
        ).strip()
        # 半角・全角カンマ、読点、空白、スラッシュ、プラス記号を区切りとして扱う。
        # 例: 1,3,5 / 1 3 5 / アイスドリンク、デザート
        parts = [
            part.strip()
            for part in re.split(r"[,，、\s/＋+]+", selected)
            if part.strip()
        ]
        selected_genres: list[str] = []
        invalid = False
        for part in parts:
            if part.isdigit() and 1 <= int(part) <= len(genres):
                genre = genres[int(part) - 1]
            elif part in genres:
                genre = part
            else:
                invalid = True
                break
            if genre not in selected_genres:
                selected_genres.append(genre)
        if selected_genres and not invalid:
            break
        print("一覧にあるジャンル名または番号を入力してください。")

    genre_counts: dict[str, int] = {}
    print("\n各ジャンルから選ぶ個数を入力してください。")
    for genre in selected_genres:
        genre_counts[genre] = ask_int(f"『{genre}』から選ぶ個数: ", minimum=0)
    if sum(genre_counts.values()) == 0:
        raise ValueError("少なくとも1つのジャンルで1個以上を指定してください。")
    budget = ask_int("割引後の合計金額の上限を入力してください（円）: ", minimum=0)

    previous_signature: tuple[tuple[str, str, int], ...] | None = None

    while True:
        selected_products = pick_products(products, genre_counts, budget, rng)

        # 可能な限り、直前とは異なる組合せを表示する。
        # 条件上ほかの組合せがない場合は、同じ結果になることがあります。
        for _ in range(20):
            signature = tuple(
                sorted((product.name, product.size, product.price) for product in selected_products)
            )
            if previous_signature is None or signature != previous_signature:
                break
            selected_products = pick_products(products, genre_counts, budget, rng)

        previous_signature = tuple(
            sorted((product.name, product.size, product.price) for product in selected_products)
        )
        print_result(selected_products, budget)

        while True:
            answer = input(
                "\n同じ条件で再抽選しますか？ [Enter / y = 再抽選, n = 終了]: "
            ).strip().lower()

            if answer in ("", "y", "yes", "はい"):
                print("\n同じ条件で再抽選します。")
                break
            if answer in ("n", "no", "いいえ"):
                print("終了します。")
                return

            print("Enter、y、またはnを入力してください。")


def print_result(products: list[Product], budget: int) -> None:
    subtotal = sum(product.price for product in products)
    discount, discount_details = calculate_discount(products)
    total = subtotal - discount

    print("\n選ばれた商品")
    for index, product in enumerate(products, start=1):
        print(f"{index:>2}. {product.label}  {product.price:,}円")
    print("-" * 40)
    print(f"通常合計: {subtotal:,}円")
    for detail in discount_details:
        print(detail)
    print(f"割引合計: -{discount:,}円")
    print(f"支払額: {total:,}円")
    print(f"残額: {budget - total:,}円")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="ジャンルごとの指定個数と予算内でドトール商品を選びます。")
    parser.add_argument("--excel", type=Path, default=Path(__file__).with_name(DEFAULT_EXCEL), help="商品一覧Excelのパス")
    parser.add_argument("--genre-count", action="append", default=[], metavar='"ジャンル=個数"', help='例: --genre-count "アイスドリンク=2"')
    parser.add_argument("--budget", type=int, help="割引後の合計金額の上限")
    parser.add_argument("--seed", type=int, help="乱数シード")
    parser.add_argument("--list-genres", action="store_true", help="利用可能なジャンル一覧を表示して終了")
    return parser

def parse_genre_counts(values: list[str]) -> dict[str, int]:
    genre_counts: dict[str, int] = {}
    for value in values:
        if "=" not in value:
            raise ValueError(f"『{value}』は「ジャンル=個数」の形式で指定してください。")
        genre, count_text = value.rsplit("=", 1)
        genre = genre.strip()
        try:
            count = int(count_text.strip())
        except ValueError as exc:
            raise ValueError(f"『{value}』の個数が整数ではありません。") from exc
        if not genre:
            raise ValueError("ジャンル名が空です。")
        if count < 0:
            raise ValueError("個数は0以上にしてください。")
        genre_counts[genre] = genre_counts.get(genre, 0) + count
    return genre_counts

def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    rng = random.Random(args.seed)
    try:
        products = load_products(args.excel)
        if args.list_genres:
            print("\n".join(available_genres(products)))
            return 0
        if args.genre_count or args.budget is not None:
            if not args.genre_count or args.budget is None:
                parser.error("--genre-countと--budgetは両方指定してください。")
            genre_counts = parse_genre_counts(args.genre_count)
            selected_products = pick_products(products, genre_counts, args.budget, rng)
            print_result(selected_products, args.budget)
        else:
            interactive(products, rng)
        return 0
    except (FileNotFoundError, ValueError) as error:
        print(f"エラー: {error}", file=sys.stderr)
        return 1

if __name__ == "__main__":
    raise SystemExit(main())
